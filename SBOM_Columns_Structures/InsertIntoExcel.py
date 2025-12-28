import openpyxl
from openpyxl.styles import Alignment,Font,Border,Side
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
import os
from ExcelRows import excelrows
import glob

def find_pdf_files(directory):
    pdf_files = glob.glob(os.path.join(directory, '*.pdf'))
    return [os.path.basename(file) for file in pdf_files]

def find_excel_files(directory):
    excel_files = glob.glob(os.path.join(directory, '*.xlsx'))
    return [os.path.basename(file) for file in excel_files]

def cell_formatting(cell):
    cell.alignment = Alignment(horizontal='center',vertical='center')
    cell.font      = Font(name="Calibri",size=14)
    border_style   = Side(border_style="thin",color="000000")
    cell.border    = Border(left=border_style,right=border_style,top=border_style,bottom=border_style)

def set_workbook_objects():
    if excelrows.get_workbook_flag()==0:
        workbooks_list = []
        directory_path = "./static/"
        pdf_files = find_pdf_files(directory_path)
        #print("lenght of pdf files:",len(pdf_files))
        for i in range(len(pdf_files)):    
            workbook_obj = Workbook()
            workbook_obj = load_workbook('data.xlsx')
            workbooks_list.append(workbook_obj)
            excelrows.set_workbooks(workbooks_list)
            excelrows.set_workbook_flag(1)
        #print(workbooks_list)

def get_workbook_objects_file_name():
    directory_path = "./static/"
    excel_files = find_excel_files(directory_path)
    wl = excelrows.get_workbooks()
    #print(wl)
    workbook = wl[len(excel_files)]
    # worksheet = workbook.worksheets[0]
    # sheet2  = workbook.worksheets[1]
    file=open("Filename.txt","r")
    static_file_name=file.readline()
    file.close()
    static_folder_excel_file = static_file_name+'.xlsx'
    downloads_folder_path = "C:/Users/Sai Pranay/Downloads"
    static_index = static_file_name.index("/")
    excel_file = downloads_folder_path + static_file_name[static_index:] +'.xlsx'
    #print(static_folder_excel_file)
    #print(excel_file)
    l = [workbook,static_folder_excel_file,excel_file]
    return l


def insert_drawing_details_into_excel(Drawing_Number,Sheet_Number,Revision_Number):
    workbook_and_file = get_workbook_objects_file_name()
    workbook = workbook_and_file[0]
    worksheet = workbook.worksheets[0] 
    sheet2 = workbook.worksheets[1]
    static_folder_excel_file = workbook_and_file[1]
    excel_file = workbook_and_file[2]
    dot_index = Drawing_Number.index(".")
    slash_index = Drawing_Number.index("/")
    project_code = Drawing_Number[slash_index+1:dot_index]
    cell_project_code = worksheet.cell(row=2,column=1,value=project_code)
    cell_formatting(cell_project_code)
    cell_drawing_number = worksheet.cell(row=4,column=2,value=Drawing_Number)
    cell_formatting(cell_drawing_number)
    cell_sheet_number = worksheet.cell(row=4,column=4,value=Sheet_Number)
    cell_formatting(cell_sheet_number)
    cell_revision_number = worksheet.cell(row=4,column=6,value=Revision_Number)
    cell_formatting(cell_revision_number)
    img = Image("static/images/shrijee1.png")
    worksheet.add_image(img,'J1')
    worksheet['J1'].alignment = Alignment(horizontal='center',vertical='center')
    workbook.save(static_folder_excel_file)
    workbook.save(excel_file)
    excelrows.set_current_row(6)
    excelrows.set_pivot_row(6)
    
def insert_cleat_details_into_excel(l):
    workbook_and_file = get_workbook_objects_file_name()
    workbook = workbook_and_file[0]
    worksheet = workbook.worksheets[0] 
    sheet2 = workbook.worksheets[1]
    static_folder_excel_file = workbook_and_file[1]
    excel_file = workbook_and_file[2]
    for i in range(len(l)):
        plate_flag = 0
        excel_l = []
        row = excelrows.get_current_row()
        excel_l.append(row - 5)
        excel_l.append(l[i].get_description())
        excel_l.append(l[i].get_section())
        excel_l.append(l[i].get_length())
        excel_l.append(l[i].get_quantity())
        if "THK" in l[i].get_section():
            excel_l.append(7.85)
            plate_flag = 1
        for col, value in enumerate(excel_l, start=1):  # Start from column 1
            cell = worksheet.cell(row=row, column=col,value=value)
            cell_formatting(cell)
        target_column = f'F'
        if "ISMB" in l[i].get_section():
            standard_weight_formula = f'=IF(COUNTIF(Sheet2!A2:A15,C${row}) > 0, INDEX(Sheet2!B2:B15, MATCH(C${row}, Sheet2!A2:A15, 0)))'
        elif "ISMC" in l[i].get_section():
            standard_weight_formula = f'=IF(COUNTIF(Sheet2!C2:C15,C${row}) > 0, INDEX(Sheet2!D2:D15, MATCH(C${row}, Sheet2!C2:C15, 0)))'
        elif "ISA" in l[i].get_section():
            standard_weight_formula = f'=IF(COUNTIF(Sheet2!E2:E139,C${row}) > 0, INDEX(Sheet2!F2:F139, MATCH(C${row}, Sheet2!E2:E139, 0)))'
        if plate_flag == 0:
            worksheet[f'{target_column}{row}'] = standard_weight_formula
            cell_formatting(worksheet[f'{target_column}{row}'])
            target_column = 'G'
            weight_formula = f'=D${row}*E${row}*F${row}/1000'
        elif plate_flag == 1:
            target_column = 'G'
            weight_formula = f'=(VALUE(LEFT(C{row}, FIND("THK", C{row})-1)) * VALUE(LEFT(D{row}, FIND("X", D{row})-1)) * RIGHT(D{row}, LEN(D{row}) - FIND("X", D{row})) * E{row} * F{row})/1000000'
        worksheet[f'{target_column}{row}'] = weight_formula
        cell_formatting(worksheet[f'{target_column}{row}'])
        #print("Row Inserted")
        excelrows.set_current_row(excelrows.get_current_row()+1)
        #workbook.save(static_folder_excel_file)

def insert_pivot(columndetails_sections):
    workbook_and_file = get_workbook_objects_file_name()
    workbook = workbook_and_file[0]
    worksheet = workbook.worksheets[0] 
    sheet2 = workbook.worksheets[1]
    static_folder_excel_file = workbook_and_file[1]
    excel_file = workbook_and_file[2]
    for cds in columndetails_sections:
        l = [cds]
        row = excelrows.get_pivot_row()
        for col, value in enumerate(l, start=8):  # Start from column 9
            cell = worksheet.cell(row=row, column=col,value = value)
            cell_formatting(cell)
        start_row = 5
        end_row = excelrows.get_current_row() - 1
        target_column = f'I'
        formula = f'=SUMPRODUCT(--(C${start_row}:C${end_row}=H{row}), D${start_row}:D${end_row}, E${start_row}:E${end_row})/1000'
        worksheet[f'{target_column}{row}'] = formula
        cell_formatting(worksheet[f'{target_column}{row}'])
        target_column = f'J'
        formula = f'=SUMIF(C${start_row}:C${end_row},H{row},G${start_row}:G${end_row})'
        worksheet[f'{target_column}{row}'] = formula
        cell_formatting(worksheet[f'{target_column}{row}'])
        excelrows.set_pivot_row(excelrows.get_pivot_row() + 1)
        #workbook.save(static_folder_excel_file)
    row = excelrows.get_pivot_row() - 1
    total_lenght_formula = f'=SUM(I6:J${row})'
    total_weight_formula = f'=SUM(J6:K${row})'
    row = row + 1
    target_column = f'H'
    worksheet[f'{target_column}{row}'] = "Total"
    cell_formatting(worksheet[f'{target_column}{row}'])
    target_column = f'I'
    worksheet[f'{target_column}{row}'] = total_lenght_formula
    cell_formatting(worksheet[f'{target_column}{row}'])
    target_column = f'J'
    worksheet[f'{target_column}{row}'] = total_weight_formula
    cell_formatting(worksheet[f'{target_column}{row}'])
    row = row + 1
    worksheet.merge_cells(start_row=row,start_column=8,end_row=row,end_column=10)
    row = row + 1
    target_column = f'I'
    worksheet[f'{target_column}{row}'] = "DESIGN APPROVED FOR MANUFACTURING"
    worksheet[f'{target_column}{row}'].font = Font(bold=True)
    cell_formatting(worksheet[f'{target_column}{row}'])
    worksheet.merge_cells(start_row=row,start_column=9,end_row=row,end_column=10)
    worksheet.merge_cells(start_row=row,start_column=8,end_row=row+6,end_column=8)
    target_column = f'H'
    border_style = Side(border_style="thin",color="000000")
    worksheet[f'{target_column}{row}'].border = Border(left=border_style,right=border_style,top=border_style,bottom=border_style)
    worksheet[f'{target_column}{row+1}'].border = Border(left=border_style)
    worksheet[f'{target_column}{row+2}'].border = Border(left=border_style)
    worksheet[f'{target_column}{row+3}'].border = Border(left=border_style)
    worksheet[f'{target_column}{row+4}'].border = Border(left=border_style)
    worksheet[f'{target_column}{row+5}'].border = Border(left=border_style)
    worksheet[f'{target_column}{row+6}'].border = Border(left=border_style,right=border_style,top=border_style,bottom=border_style)
    target_column = f'I'
    row = row + 1
    worksheet[f'{target_column}{row}'] = "SIGN"
    worksheet[f'{target_column}{row}'].font = Font(bold=True)
    worksheet[f'{target_column}{row}'].font = Font(name="Calibri",size=14)
    worksheet[f'{target_column}{row}'].alignment = Alignment(horizontal='center',vertical="top")
    border_style   = Side(border_style="thin",color="000000")
    worksheet[f'{target_column}{row}'].border = Border(left=border_style,right=border_style,top=border_style,bottom=border_style)
    worksheet.merge_cells(start_row=row,start_column=9,end_row=row+5,end_column=9)
    target_column = f'J'
    worksheet[f'{target_column}{row}'] = "DATE"
    worksheet[f'{target_column}{row}'].font = Font(bold=True)
    worksheet[f'{target_column}{row}'].font = Font(name="Calibri",size=14)
    worksheet[f'{target_column}{row}'].alignment = Alignment(horizontal='center',vertical="top")
    border_style   = Side(border_style="thin",color="000000")
    worksheet[f'{target_column}{row}'].border = Border(left=border_style,right=border_style,top=border_style,bottom=border_style)
    worksheet.merge_cells(start_row=row,start_column=10,end_row=row+5,end_column=10)
    #workbook.save(static_folder_excel_file)
    #workbook.save(excel_file)
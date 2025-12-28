try:
    import openpyxl
    from openpyxl.styles import Alignment,Font,Border,Side
    from openpyxl.drawing.image import Image
    from PIL import Image as PILImage
    from openpyxl.utils import get_column_letter
    from openpyxl import Workbook, load_workbook
    import os
    from ExcelRows import excelrows
    import glob

    def find_pdf_files(directory):
        pdf_files = glob.glob(os.path.join(directory, '*.pdf'))
        return [os.path.basename(file) for file in pdf_files]

    def find_excel_files(directory):
        excel_files = glob.glob(os.path.join(directory, '*.xlsx'))
        return [os.path.basename(file) for file in excel_files]

    def get_file_names():
        file = open("FileName.txt","r")
        name = file.readline()
        file.close() 
        static_excel = name + '.xlsx'
        slash_index = name.index("static/")
        slash_index = slash_index + 7
        name = name[slash_index:]
        name = "C:/Users/Sai Pranay/Downloads/" + name
        excel_file = name + '.xlsx'
        files_list = []
        files_list.append(excel_file)
        files_list.append(static_excel)
        return files_list

    def formatting_data(cell):
        cell.alignment = Alignment(horizontal='center',vertical='center')
        cell.font      = Font(name="Calibri",size=14)
        border_style   = Side(border_style="thin",color="000000")
        cell.border    = Border(left=border_style,right=border_style,top=border_style,bottom=border_style)
        files_names = get_file_names()
        #workbook.save(files_names[0])
        #workbook.save(files_names[1])
    
    def insert_drawing_details(Drawing_Number,Sheet_Number,Revision_Number):
        if excelrows.get_workbook_flag()==0:
            workbooks_list = []
            directory_path = "./static/"
            pdf_files = find_pdf_files(directory_path)
            for i in range(len(pdf_files)):    
                workbook_obj = Workbook()
                workbook_obj = load_workbook('data.xlsx')
                workbooks_list.append(workbook_obj)
                excelrows.set_workbooks(workbooks_list)
                excelrows.set_workbook_flag(1)
        directory_path = "./static/"
        excel_files = find_excel_files(directory_path)
        wl = excelrows.get_workbooks()
        workbook = wl[len(excel_files)]
        worksheet = workbook.worksheets[0]
        sheet  = workbook.worksheets[1]
        dot_index = Drawing_Number.index(".")
        slash_index = Drawing_Number.index("/")
        project_code = Drawing_Number[slash_index+1:dot_index]
        worksheet.merge_cells(start_row =1,start_column=1,end_row=1,end_column=2)
        worksheet['A1'].value = "PROJECT CODE"
        worksheet['A1'].alignment = Alignment(horizontal='center',vertical='center')
        worksheet['A1'].font = Font(name="Calibri",size=14,bold=True)
        border_style   = Side(border_style="thin",color="000000")
        worksheet['A1'].border = Border(left=border_style,right=border_style,top=border_style)
        worksheet.merge_cells(start_row =2,start_column=1,end_row=2,end_column=2)
        worksheet['A2'].value = project_code
        worksheet['A2'].alignment = Alignment(horizontal='center',vertical='center')
        worksheet['A2'].font = Font(name="Calibri",size=14,bold=True)
        border_style   = Side(border_style="thin",color="000000")
        worksheet['A2'].border = Border(left=border_style,right=border_style,bottom=border_style)
        worksheet.merge_cells("J1:J4")
        img = Image("static/images/shrijee1.png")
        worksheet.add_image(img,'J1')
        worksheet.merge_cells(start_row=1,start_column=10,end_row=4,end_column=10)
        worksheet['J1'].alignment = Alignment(horizontal='center',vertical='center')
        l = []
        row = excelrows.get_drawingdetails_row()
        l.append(Drawing_Number)          
        l.append(Sheet_Number)            
        l.append(Revision_Number)
        worksheet['B4'].value = Drawing_Number
        worksheet['B4'].alignment = Alignment(horizontal='center',vertical='center')
        worksheet['D4'].value = Sheet_Number
        worksheet['D4'].alignment = Alignment(horizontal='center',vertical='center')
        worksheet['F4'].value = Revision_Number
        worksheet['F4'].alignment = Alignment(horizontal='center',vertical='center')
        files_names = get_file_names()
        # workbook.save(files_names[0])
        # workbook.save(files_names[1])
        excelrows.set_current_row(6)
        excelrows.set_pivot_row(6)
    
    def insert_beam_into_excel(sr_no,mkd,ism):
        directory_path = "./static/"
        excel_files = find_excel_files(directory_path)
        wl = excelrows.get_workbooks()
        workbook = wl[len(excel_files)]
        worksheet = workbook.worksheets[0]
        sheet  = workbook.worksheets[1]
        l = []
        l.append(sr_no)                                          
        l.append(mkd)
        l.append(ism.get_beam_section())
        l.append(ism.get_beam_length())
        l.append(ism.get_beam_quantity())
        row = excelrows.get_current_row() 
        for col, value in enumerate(l, start=1):  # Start from column 1
            cell = worksheet.cell(row=row, column=col,value=value)
            formatting_data(cell)
        target_column = 'F'
        if "ISMB" in ism.get_beam_section():
            standard_weight_formula = f'=IF(COUNTIF(Sheet2!A2:A15,C${row}) > 0, INDEX(Sheet2!B2:B15, MATCH(C${row}, Sheet2!A2:A15, 0)))'
        if "ISMC" in ism.get_beam_section():
            standard_weight_formula = f'=IF(COUNTIF(Sheet2!C2:C15,C${row}) > 0, INDEX(Sheet2!D2:D15, MATCH(C${row}, Sheet2!C2:C15, 0)))'
        worksheet[f'{target_column}{row}'] = standard_weight_formula
        formatting_data(worksheet[f'{target_column}{row}'])
        target_column = 'G'
        weight_formula = f'=D${row}*E${row}*F${row}/1000'
        worksheet[f'{target_column}{row}'] = weight_formula
        formatting_data(worksheet[f'{target_column}{row}'])
        excelrows.set_current_row(excelrows.get_current_row() + 1)
        files_names = get_file_names()
        #workbook.save(files_names[0])
        #workbook.save(files_names[1])
    
    def insert_angle_cleat_into_excel(sr_no,mkd,isa):
        directory_path = "./static/"
        excel_files = find_excel_files(directory_path)
        wl = excelrows.get_workbooks()
        workbook = wl[len(excel_files)]
        worksheet = workbook.worksheets[0]
        sheet  = workbook.worksheets[1]
        l = []
        l.append(sr_no)                                  
        l.append(mkd)
        isa_section = isa.get_angle_cleat_section()
        count_x = isa_section.count('x')
        if count_x < 2:
            isa_section = isa_section.replace('x','')
            for i in range(1,139):
                if isa_section == sheet['G'+str(i)].value:
                    isa_section = sheet['E'+str(i)].value
                    break
        l.append(isa_section)
        l.append(isa.get_angle_cleat_length())
        l.append(isa.get_angle_cleat_quantity())
        row = excelrows.get_current_row()
        for col, value in enumerate(l, start=1):  # Start from column 1
            cell = worksheet.cell(row=row, column=col, value=value)
            formatting_data(cell)
        target_column = 'F'
        standard_weight_formula = f'=IF(COUNTIF(Sheet2!E2:E139,C${row}) > 0, INDEX(Sheet2!F2:F139, MATCH(C${row}, Sheet2!E2:E139, 0)))'
        worksheet[f'{target_column}{row}'] = standard_weight_formula
        formatting_data(worksheet[f'{target_column}{row}'])
        target_column = 'G'
        weight_formula = f'=D${row}*E${row}*F${row}/1000'
        worksheet[f'{target_column}{row}'] = weight_formula
        formatting_data(worksheet[f'{target_column}{row}'])
        excelrows.set_current_row(excelrows.get_current_row() + 1)
        files_names = get_file_names()
        #workbook.save(files_names[0])
        #workbook.save(files_names[1])
    
    def insert_plate_into_excel(sr_no,mkd,plt):
        directory_path = "./static/"
        excel_files = find_excel_files(directory_path)
        wl = excelrows.get_workbooks()
        workbook = wl[len(excel_files)]
        worksheet = workbook.worksheets[0]
        sheet  = workbook.worksheets[1]
        l = []
        l.append(sr_no)                                  
        l.append(mkd)
        l.append(plt.get_plate_thickness())
        l.append(plt.get_plate_dimensions())
        l.append(plt.get_plate_quantity())
        l.append(plt.get_plate_standard_weight())
        row = excelrows.get_current_row()
        for col, value in enumerate(l, start=1):  # Start from column 1
            cell = worksheet.cell(row=row, column=col, value=value)
            formatting_data(cell)
        target_column = 'G'
        weight_formula = f'=(VALUE(LEFT(C{row}, FIND("THK", C{row})-1)) * VALUE(LEFT(D{row}, FIND("x", D{row})-1)) * RIGHT(D{row}, LEN(D{row}) - FIND("x", D{row})) * E{row} * F{row})/1000000'
        worksheet[f'{target_column}{row}'] = weight_formula
        formatting_data(worksheet[f'{target_column}{row}'])
        excelrows.set_current_row(excelrows.get_current_row() + 1)
        files_names = get_file_names()
        #workbook.save(files_names[0])
        #workbook.save(files_names[1])
    
    def insert_pivot(beams_angles_plates):
        directory_path = "./static/"
        excel_files = find_excel_files(directory_path)
        wl = excelrows.get_workbooks()
        workbook = wl[len(excel_files)]
        worksheet = workbook.worksheets[0]
        sheet  = workbook.worksheets[1]
        for bap in beams_angles_plates:
            l = [bap]
            row = excelrows.get_pivot_row()
            for col, value in enumerate(l, start=8):  # Start from column 8
                cell = worksheet.cell(row=row, column=col, value = value)
                formatting_data(cell)
            start_row = 2
            end_row = excelrows.get_current_row() - 1
            target_column = f'I'
            formula = f'=SUMPRODUCT(--(C${start_row}:C${end_row}=H{row}), D${start_row}:D${end_row}, E${start_row}:E${end_row})/1000'
            worksheet[f'{target_column}{row}'] = formula
            formatting_data(worksheet[f'{target_column}{row}'])
            target_column = f'J'
            formula = f'=SUMIF(C${start_row}:C${end_row},H{row},G${start_row}:G${end_row})'
            worksheet[f'{target_column}{row}'] = formula
            formatting_data(worksheet[f'{target_column}{row}'])
            excelrows.set_pivot_row(excelrows.get_pivot_row() + 1)
            files_names = get_file_names()
            #workbook.save(files_names[0])
            #workbook.save(files_names[1])
        
    def insert_pivot_total():
        directory_path = "./static/"
        excel_files = find_excel_files(directory_path)
        wl = excelrows.get_workbooks()
        workbook = wl[len(excel_files)]
        worksheet = workbook.worksheets[0]
        sheet  = workbook.worksheets[1]
        row = excelrows.get_pivot_row() 
        total_lenght_formula = f'=SUM(I6:I${row-1})'
        total_weight_formula = f'=SUM(J6:J${row-1})'
        target_column = f'H'
        worksheet[f'{target_column}{row}'] = "Total"
        formatting_data(worksheet[f'{target_column}{row}'])
        target_column = f'I'
        worksheet[f'{target_column}{row}'] = total_lenght_formula
        formatting_data(worksheet[f'{target_column}{row}'])
        target_column = f'J'
        worksheet[f'{target_column}{row}'] = total_weight_formula
        formatting_data(worksheet[f'{target_column}{row}'])
        files_names = get_file_names()
        workbook.save(files_names[0])
        workbook.save(files_names[1])
        excelrows.set_pivot_row(excelrows.get_pivot_row() + 1)
        row = excelrows.get_pivot_row()
        worksheet.merge_cells(start_row=row,start_column=8,end_row=row,end_column=10)
        row = row + 1
        target_column = f'I'
        worksheet[f'{target_column}{row}'] = "DESIGN APPROVED FOR MANUFACTURING"
        worksheet[f'{target_column}{row}'].font = Font(bold=True)
        formatting_data(worksheet[f'{target_column}{row}'])
        # workbook.save(excel_file)
        worksheet.merge_cells(start_row=row,start_column=9,end_row=row,end_column=10)
        worksheet.merge_cells(start_row=row,start_column=8,end_row=row+6,end_column=8)
        target_column = f'H'
        border_style = Side(border_style="thin",color="000000")
        worksheet[f'{target_column}{row}'].border = Border(left=border_style,right=border_style,top=border_style,bottom=border_style)
        worksheet[f'{target_column}{row+1}'].border = Border(left=border_style)
        worksheet[f'{target_column}{row+2}'].border = Border(left=border_style)
        worksheet[f'{target_column}{row+3}'].border = Border(left=border_style)
        worksheet[f'{target_column}{row+4}'].border = Border(left=border_style)
        worksheet[f'{target_column}{row+5}'].border = Border(left=border_style)
        worksheet[f'{target_column}{row+6}'].border = Border(left=border_style,right=border_style,top=border_style,bottom=border_style)
        target_column = f'I'
        row = row + 1
        worksheet[f'{target_column}{row}'] = "SIGN"
        worksheet[f'{target_column}{row}'].font = Font(bold=True)
        worksheet[f'{target_column}{row}'].font = Font(name="Calibri",size=14)
        worksheet[f'{target_column}{row}'].alignment = Alignment(horizontal='center',vertical="top")
        border_style   = Side(border_style="thin",color="000000")
        worksheet[f'{target_column}{row}'].border = Border(left=border_style,right=border_style,top=border_style,bottom=border_style)
        worksheet.merge_cells(start_row=row,start_column=9,end_row=row+5,end_column=9)
        target_column = f'J'
        worksheet[f'{target_column}{row}'] = "DATE"
        worksheet[f'{target_column}{row}'].font = Font(bold=True)
        worksheet[f'{target_column}{row}'].font = Font(name="Calibri",size=14)
        worksheet[f'{target_column}{row}'].alignment = Alignment(horizontal='center',vertical="top")
        border_style   = Side(border_style="thin",color="000000")
        worksheet[f'{target_column}{row}'].border = Border(left=border_style,right=border_style,top=border_style,bottom=border_style)
        worksheet.merge_cells(start_row=row,start_column=10,end_row=row+5,end_column=10)
        #workbook.save(excel_file) 
        files_names = get_file_names()
        workbook.save(files_names[0])
        workbook.save(files_names[1])

except Exception as e:
    print(e)
    print("ERROR OCCURED in text_recogntion")
    file = open("FileName.txt","r")
    name = file.readline()
    file.close()      
    os.remove(name+'.png')
    